import os
import openpyxl
from globals import dir_global


class ExcelReader:
    def __init__(self, excel_path):
        self.excel_path = os.path.join(dir_global.DATA_FILES_PATH, excel_path)

    def get_row_count(self, sheet_name):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    def get_column_count(self, sheet_name):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[sheet_name]
        return sheet.max_column

    def get_cell_data(self, sheet_name, row_number, column_number):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_number, column=column_number).value

    def set_cell_data(self, sheet_name, row_number, column_number, data):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.excel_path)

    def get_data_from_excel(self, sheet_name):
        final_list = []
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[sheet_name]
        total_rows = sheet.max_row
        total_cols = sheet.max_column

        for r in range(2, total_rows + 1):
            row_list = []
            for c in range(1, total_cols + 1):
                row_list.append(sheet.cell(row=r, column=c).value)
            final_list.append(row_list)

        return final_list
